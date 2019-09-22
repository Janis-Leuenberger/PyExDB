from openpyxl import *
from prettytable import PrettyTable
import sys

def BasicAllocation(BaseCommand):
    #Help
    if BaseCommand[0] == "HELP":
        print("--------------------------------------------------------------------------------------")
        print("Create a database:   'CREATE DATABASE database_name'")
        print("Use a database:      'USE DATABASE database_name'")
        print("Create a table:      'CREATE TABLE table_name ( column_name column_name column_name )'")
        print("Do an insert:        'INSERT INTO table_name VALUES ( value value value )'")
        print("Do a select:         'SELECT column_name FROM table_name'")
        print("Use the exact Syntax!!!")
        print("--------------------------------------------------------------------------------------")

    #Exit
    elif BaseCommand[0] == "EXIT":
        sys.exit()
    #Create
    elif (BaseCommand[0] == "CREATE"):
        if (BaseCommand[1] == "DATABASE"):
            CreateDatabase(BaseCommand[2])
        elif(BaseCommand[1] == "TABLE"):
            CreateTable(BaseCommand)
    #Insert
    elif (BaseCommand[0] == "INSERT"):
        if (BaseCommand[1] == "INTO" and BaseCommand[3] == "VALUES"):
            DoInsertInto(BaseCommand)
    #Select
    elif (BaseCommand[0] == "SELECT"):
        DoSelectFrom(BaseCommand)
    #Use
    elif (BaseCommand[0] == "USE"):
        if (BaseCommand[1] == "DATABASE"):
            DoUseDatabase(BaseCommand[2])



#CreateDatabase
def CreateDatabase(DataBaseName):
    try:
        wb = Workbook()
        wb.save(filename='../database/' + DataBaseName + '.xlsx')
        print("Created database '" + DataBaseName + "'.")
        DoUseDatabase(DataBaseName)
    except:
        print("Could not create database '" + DataBaseName + "'.")


def CreateTable(BaseCommand):
    try:
        db = load_workbook(filename='../database/' + GetDatabaseName() + ".xlsx")
        table = db.create_sheet(BaseCommand[2])
        for i in range(len(BaseCommand) - 5):
            spot = TranslateToAlphabet(i) + str(1)
            table[spot] = str(BaseCommand[i + 4])
        db.save(filename='../database/' + GetDatabaseName() + ".xlsx")
        print("Created table '" + BaseCommand[2] + "'.")
    except:
        print("Could no create table '" + BaseCommand[2] + "'.")


def DoInsertInto(BaseCommand):
    try:
        db = load_workbook(filename='../database/' + GetDatabaseName() + ".xlsx")
        table = db.get_sheet_by_name(BaseCommand[2])

        y = GetHeight(table)
        i = 5
        while i < len(BaseCommand) - 1:
            a = i - 5
            table[TranslateToAlphabet(a) + str(y)] = BaseCommand[i]
            i += 1
        db.save(filename='../database/' + GetDatabaseName() + ".xlsx")
        print("Insert was successful.")
    except:
        print("Insert failed.")



def DoSelectFrom(BaseCommand):
    try:
        db = load_workbook(filename='../database/' + GetDatabaseName() + ".xlsx")
        index = BaseCommand.index("FROM")
        tableindex = index + 1
        table = db.get_sheet_by_name(BaseCommand[tableindex])

        alist = []
        if BaseCommand[1] == "*":
            k = 0
            while k < GetHeight(table) - 1:
                k += 1
                i = -1
                while i < GetWidth(table) - 1:
                    i += 1
                    spot = TranslateToAlphabet(i) + str(k)
                    alist.append(table[spot].value)
            width = GetWidth(table)
            PrintArray(alist, width)
            return
        else:
            alist = []
            for item in BaseCommand:
                i = 0
                while i < GetWidth(table):
                    spot = TranslateToAlphabet(i) + str(1)
                    if(item != table[spot].value):
                        i += 1
                        continue

                    k = 0
                    while k < GetHeight(table) - 1:
                        k += 1
                        spot = TranslateToAlphabet(i) + str(k)
                        alist.append(table[spot].value)
                    i += 1
            width = index - 1
            newlist = ChangeArray(alist, width)
            PrintArray(newlist, width)



    except:
        print("Select failed.")


def DoUseDatabase(DataBaseName):
    try:
        file = open("DBName.txt", "w")
        file.write(DataBaseName)
        file.close()
        print("Changed database to '" + DataBaseName + "'.")
    except:
        print("Could not change database.")


#Get DataBaseName
def GetDatabaseName():
    file = open("DBName.txt", "r")
    text = file.read()
    return text

#other stuff
def TranslateToAlphabet(number):
    number += 66
    return chr(number)

#Get Height
def GetHeight(table):
    i = 0
    while True:
        i += 1
        spot = "B" + str(i)
        value = table[spot].value
        if str(value) == "None":
            return i

#Get Width
def GetWidth(table):
    i = 0
    while True:
        i += 1
        char = TranslateToAlphabet(i)
        spot = char + "1"
        value = table[spot].value
        if str(value) == "None":
            return i
    return

def ChangeArray(alist, width):
    numelements = len(alist) / width
    newlist = []

    i = 0
    i2 = 0
    while i < len(alist):
        k = 0
        while k < width:
            spot = i2 + int(numelements * k)
            newlist.append(alist[int(spot)])
            k += 1
        i += width
        i2 += 1

    return newlist

def PrintArray(alist, width):
    if width > 5:
        print("The use of more then 5 columns is not supported atm.")
        return

    if width == 1:
        t = PrettyTable([alist[0]])
        i = 1
        while i < len(alist):
            t.add_row([alist[i]])
            i += 1
    elif width == 2:
        t = PrettyTable([alist[0], alist[1]])
        i = 2
        while i < len(alist):
            t.add_row([alist[i], alist[i + 1]])
            i += 2
    elif width == 3:
        t = PrettyTable([alist[0], alist[1], alist[2]])
        i = 3
        while i < len(alist):
            t.add_row([alist[i], alist[i + 1], alist[i + 2]])
            i += 3
    elif width == 4:
        t = PrettyTable([alist[0], alist[1], alist[2], alist[3]])
        i = 4
        while i < len(alist):
            t.add_row([alist[i], alist[i + 1], alist[i + 2], alist[i + 3]])
            i += 4
    elif width == 5:
        t = PrettyTable([alist[0], alist[1], alist[2], alist[3], alist[4]])
        i = 5
        while i < len(alist):
            t.add_row([alist[i], alist[i + 1], alist[i + 2], alist[i + 3], alist[i + 3]])
            i += 5

    print(t)



